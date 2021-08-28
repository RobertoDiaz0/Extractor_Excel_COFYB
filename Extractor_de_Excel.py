import  tkinter as tk
import openpyxl
import os
import re
Carpeta_actual = os.getcwd()
Carpeta_actual = Carpeta_actual + "/"
Nombre_del_archivo = ""
Nombre_ingresado = ""
Nombre_del_archivo = ""
Archivo_colegiados = Carpeta_actual + "Colegiados.xlsx"
Archivos_de_la_carpeta = os.listdir()
Tipo_de_error = 3
Mensajes = ('4. El proceso ha concluído exitosamente', 'Error 2: No se encuentra archivo con el nombre ingresado', 'Error 1: No se encuentra el archivo\n"Colegiados.xlsx"\nen esta carpeta', 'Error 1: No se encuentra el archivo"Colegiados.xlsx"\nen esta carpeta\n\nError 2: No se encuentra archivo con el nombre ingresado')
Archivos_xlsx = []
Contador_error_1 = 0
Contador_error_2 = 0
Progreso = "0"
Mensaje_final = ""

def Ingresar():
    global Progreso
    Etiqueta.configure(text="\n\nAguarde mientras se procesan los datos\n\n%"+ Progreso+"\n")
    Etiqueta.update()
    global Carpeta_actual
    global Archivo_colegiados
    global Nombre_ingresado
    global Archivos_xlsx
    global Archivos_de_la_carpeta
    global Contador_error_1
    global Contador_error_2
    global Mensajes
    global Mensaje_final
    global Tipo_de_error
    global Nombre_del_archivo
    Nombre_ingresado = Cuadro_clave.get()
    for i in Archivos_de_la_carpeta:
        Resultado = re.findall(r"^.*xlsx$", i)
        if len(Resultado) > 0:
            Archivos_xlsx = Archivos_xlsx + (Resultado)
    for i in Archivos_xlsx:
        if i == Nombre_ingresado:
            Contador_error_2 = 1
        if i == "Colegiados.xlsx":
            Contador_error_1 = 2
    Tipo_de_error = Tipo_de_error - Contador_error_1
    Tipo_de_error = Tipo_de_error - Contador_error_2
    if Tipo_de_error == 0:
        Nombre_del_archivo = Carpeta_actual + Nombre_ingresado
        libro_1 = openpyxl.load_workbook(Nombre_del_archivo)
        libro_2 = openpyxl.load_workbook(Archivo_colegiados)
        hoja_1 = libro_1.active
        hoja_2 = libro_2.active
        cantidad_de_filas_1 = hoja_1.max_row
        cantidad_de_filas_2 = hoja_2.max_row
        for x in range (2, cantidad_de_filas_1):
            Progreso = x / cantidad_de_filas_1 * 100
            Progreso = int(Progreso)
            Progreso = str(Progreso)
            Etiqueta.configure(text="\n\nAguarde mientras se procesan los datos\n\n%"+ Progreso+"\n")
            Etiqueta.update()
            if hoja_1.cell(row=x,column=6).value == None:
                pass
            elif hoja_1.cell(row=x,column=6).value > 0:
                celda = hoja_1.cell(row=x,column=3).value
                cuit_cuil = ("")
                for i in celda:
                    if i.isdigit():
                        cuit_cuil += i             
                hoja_1.cell(row=x,column=9).value = cuit_cuil
# Comparo el número extraído con el CUIT de la planilla de colegiados.
# Si coinciden copio el número y nombre del colegiado.
                for i in range(2, cantidad_de_filas_2):
                    primer_número = hoja_1.cell(row = x, column = 9).value
                    segundo_número = hoja_2.cell(row = i, column = 2).value
                    if primer_número == segundo_número:
                        hoja_1.cell(row = x, column = 10).value = hoja_2.cell(row = i, column = 1).value
                        hoja_1.cell(row = x, column = 11).value = hoja_2.cell(row = i, column = 3).value
# Grabo Modificaciones
        libro_1.save(filename = Nombre_del_archivo)
    Ventana_2 = tk.Toplevel(Ventana)
    Ventana_2.configure(background='black')
    Ventana_2.geometry("420x160")
    Mensaje_final = Mensajes[Tipo_de_error]
    Etiqueta_2 = tk.Label(Ventana_2, text="\n\n"+Mensaje_final+"\n\n", fg='white', bg='black')
    Aceptar = tk.Button(Ventana_2, text="Aceptar", command=Ventana.destroy, fg='white', bg='grey')
    Etiqueta_2.pack()
    Aceptar.pack()
    
#Ventana
Ventana = tk.Tk()
Ventana.configure(background='black')
Ventana.geometry("470x240")
Etiqueta = tk.Label(Ventana, text="\n\n1. Ingrese el nombre del archivo del que quiere extraer Cuit/Cuil.\n\n2. Recuerde escribirlo con la extensión .xlsx\n\n3. El archivo debe estar en la misma carpeta que este programa\n al igual que el archivo ""'Colegiados.xlsx'""\n",fg='lightgrey', bg='black')
Btn = tk.Button(Ventana, text="Ingresar", command=Ingresar, fg='white', bg='grey')
Cuadro_clave = tk.Entry(Ventana, width=25)
Cuadro_clave.pack()
Etiqueta.pack()
Btn.pack()

Ventana.mainloop()
